import openpyxl

# Load the Excel workbook for input table
workbook = openpyxl.load_workbook('User Statistics.xlsx')

# Select the worksheet for input table
worksheet = workbook['Sheet1']

# Define empty lists for columns
names = []
uids = []
statements = []
reasons = []

# Loop through rows and append values to columns
for row in worksheet.iter_rows(min_row=2, values_only=True):
    # Replace empty cells with 0
    row = [0 if x is None else x for x in row]
    names.append(row[1])
    uids.append(row[2])
    statements.append(row[3])
    reasons.append(row[4])

# Calculate total statements and reasons for each user
total_statements = [statements[i]+reasons[i] for i in range(len(statements))]

# Sort users by total statements and reasons
sorted_indices = sorted(range(len(total_statements)), key=lambda k: (-total_statements[k], -reasons[k]))
sorted_names = [names[i] for i in sorted_indices]
sorted_uids = [uids[i] for i in sorted_indices]
sorted_total_statements = [total_statements[i] for i in sorted_indices]
sorted_statements = [statements[i] for i in sorted_indices]
sorted_reasons = [reasons[i] for i in sorted_indices]

# Create a new Excel workbook and worksheet for output
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active

# Populate the output worksheet with data
output_worksheet.cell(row=1, column=1, value="Rank")
output_worksheet.cell(row=1, column=2, value="Name")
output_worksheet.cell(row=1, column=3, value="UID")
output_worksheet.cell(row=1, column=4, value="No. of Statements")
output_worksheet.cell(row=1, column=5, value="No. of Reasons")

for i, name in enumerate(sorted_names):
    output_worksheet.cell(row=i+2, column=1, value=i+1)
    output_worksheet.cell(row=i+2, column=2, value=name)
    output_worksheet.cell(row=i+2, column=3, value=sorted_uids[i])
    output_worksheet.cell(row=i+2, column=4, value=sorted_statements[i])
    output_worksheet.cell(row=i+2, column=5, value=sorted_reasons[i])

# Save the output workbook to a file
output_workbook.save('user_rankings.xlsx')
