import openpyxl

# Load the Excel workbooks for input tables
workbook1 = openpyxl.load_workbook('userids.xlsx')
workbook2 = openpyxl.load_workbook('User Statistics.xlsx')
workbook3 = openpyxl.load_workbook('Statements Matrix.xlsx')

# Select the worksheets for input tables
worksheet1 = workbook1['Sheet1']
worksheet2 = workbook2['Sheet1']
worksheet3 = workbook3['Sheet1']

# Define empty lists for columns
teams = []
statements = []
reasons = []

# Loop through rows and append values to columns
for row in worksheet1.iter_rows(min_row=2, values_only=True):
    teams.append(row[2])

for row in worksheet2.iter_rows(min_row=2, values_only=True):
    # Replace empty cells with 0
    row = [0 if x is None else x for x in row]
    statements.append(row[3])
    reasons.append(row[4])

# Get unique team names
unique_teams = sorted(list(set(teams)))

# Calculate average statements and reasons per team
avg_statements = []
avg_reasons = []
for team in unique_teams:
    indices = [i for i, x in enumerate(teams) if x == team]
    total_statements = sum([statements[i] for i in indices])
    total_reasons = sum([reasons[i] for i in indices])
    num_users = len(indices)
    avg_statements.append(round(total_statements / num_users, 2))
    avg_reasons.append(round(total_reasons / num_users, 2))

# Sort teams by average statements and reasons
sorted_indices = sorted(range(len(avg_statements)), key=lambda k: (-avg_statements[k], -avg_reasons[k]))
sorted_teams = [unique_teams[i] for i in sorted_indices]
sorted_avg_statements = [round(avg_statements[i], 2) for i in sorted_indices]
sorted_avg_reasons = [round(avg_reasons[i], 2) for i in sorted_indices]

# Create a new Excel workbook and worksheet for output
output_workbook = openpyxl.Workbook()
output_worksheet = output_workbook.active

# Populate the output worksheet with data
output_worksheet.cell(row=1, column=1, value="Team Rank")
output_worksheet.cell(row=1, column=2, value="Thinking Teams Leaderboard")
output_worksheet.cell(row=1, column=3, value="Average Statements")
output_worksheet.cell(row=1, column=4, value="Average Reasons")

for i, team in enumerate(sorted_teams):
    output_worksheet.cell(row=i+2, column=1, value=i+1)
    output_worksheet.cell(row=i+2, column=2, value=team)
    output_worksheet.cell(row=i+2, column=3, value=sorted_avg_statements[i])
    output_worksheet.cell(row=i+2, column=4, value=sorted_avg_reasons[i])

# Save the output workbook to a file
output_workbook.save('team_rankings.xlsx')
